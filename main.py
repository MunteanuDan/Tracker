from flask import Flask, render_template, request, redirect, url_for, flash, session
import sqlite3
from datetime import datetime, date, timedelta

from openpyxl.workbook import Workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'


# Setup database
def get_db_connection():
    conn = sqlite3.connect('pontaj.db')
    conn.row_factory = sqlite3.Row
    return conn


# Initializare baza de date
with app.app_context():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS pontaj
                 (id INTEGER PRIMARY KEY, nume TEXT, data TEXT, ora_intrare TEXT, ora_iesire TEXT, departament TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (username TEXT, password TEXT, role TEXT, departament TEXT, PRIMARY KEY(username, departament))''')
    conn.commit()
    conn.close()


# Functie pentru validarea formatului de ora
def validate_time_format(time_str):
    try:
        datetime.strptime(time_str, '%H:%M')
        return True
    except ValueError:
        return False


@app.route('/')
def index():
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        departament = request.form['departament']
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT role FROM users WHERE username = ? AND password = ? AND departament = ?",
                  (username, password, departament))
        result = c.fetchone()
        conn.close()
        if result:
            session['username'] = username
            session['departament'] = departament
            session['role'] = result['role']
            if result['role'] == 'admin':
                return redirect(url_for('admin'))
            elif result['role'] == 'viewer':
                return redirect(url_for('viewer'))
        else:
            flash("Nume de utilizator, parolă sau departament incorect")
    return render_template('login.html')


@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if 'username' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    if request.method == 'POST':
        data = request.form['data']
        ora_intrare = request.form['ora_intrare']
        ora_iesire = request.form['ora_iesire']
        concediu = 'concediu' in request.form
        username = session['username']
        departament = session['departament']

        if concediu:
            ora_intrare = 'CO'
            ora_iesire = 'CO'

        if not concediu and not (validate_time_format(ora_intrare) and validate_time_format(ora_iesire)):
            flash("Formatul orei trebuie să fie HH:MM.")
            return redirect(url_for('admin'))

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM pontaj WHERE nume = ? AND data = ? AND departament = ?",
                  (username, data, departament))
        if c.fetchone()[0] > 0:
            flash("Există deja o înregistrare pentru această zi.")
            conn.close()
            return redirect(url_for('admin'))

        c.execute("INSERT INTO pontaj (nume, data, ora_intrare, ora_iesire, departament) VALUES (?, ?, ?, ?, ?)",
                  (username, data, ora_intrare, ora_iesire, departament))
        conn.commit()
        conn.close()
        flash("Pontaj adăugat cu succes!")

    return render_template('admin.html')


@app.route('/viewer')
def viewer():
    if 'username' not in session or session['role'] != 'viewer':
        return redirect(url_for('login'))

    departament = session['departament']
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT nume FROM pontaj WHERE departament = ? ORDER BY nume", (departament,))
    users = c.fetchall()
    conn.close()
    return render_template('viewer.html', users=users)


@app.route('/vizualizeaza_pontaj/<username>')
def vizualizeaza_pontaj(username):
    if 'username' not in session or session['role'] != 'viewer':
        return redirect(url_for('login'))

    departament = session['departament']
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT nume, data, ora_intrare, ora_iesire FROM pontaj WHERE nume = ? AND departament = ? ORDER BY data",
              (username, departament))
    records = c.fetchall()
    conn.close()
    return render_template('vizualizeaza_pontaj.html', records=records)


@app.route('/creaza_cont', methods=['GET', 'POST'])
def creaza_cont():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        departament = request.form['departament']

        conn = get_db_connection()
        c = conn.cursor()
        if role == "viewer":
            c.execute("SELECT COUNT(*) FROM users WHERE role = ? AND departament = ?", (role, departament))
            viewer_count = c.fetchone()[0]
            if viewer_count >= 1:
                flash("Nu poți crea mai mult de un cont de vizualizare per departament.")
                conn.close()
                return redirect(url_for('creaza_cont'))

        c.execute("INSERT INTO users (username, password, role, departament) VALUES (?, ?, ?, ?)",
                  (username, password, role, departament))
        conn.commit()
        conn.close()
        flash(f"Cont {role} creat cu succes în departamentul {departament}!")
        return redirect(url_for('login'))

    return render_template('creaza_cont.html')


@app.route('/sterge_cont', methods=['GET', 'POST'])
def sterge_cont():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        departament = request.form['departament']

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username = ? AND password = ? AND departament = ?",
                  (username, password, departament))
        result = c.fetchone()
        if result:
            c.execute("DELETE FROM users WHERE username = ? AND departament = ?", (username, departament))
            conn.commit()
            conn.close()
            flash("Contul a fost șters cu succes!")
        else:
            conn.close()
            flash("Nume de utilizator, parolă sau departament incorect")
        return redirect(url_for('login'))

    return render_template('sterge_cont.html')


@app.route('/curata_pontaj_db')
def curata_pontaj_db():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM pontaj")
    c.execute("DELETE FROM users")
    conn.commit()
    conn.close()
    flash("Toate înregistrările au fost șterse!")
    return redirect(url_for('login'))


@app.route('/exporta_in_excel')
def exporta_in_excel():
    departament = session.get('departament')
    if not departament:
        return redirect(url_for('login'))

    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT nume FROM pontaj WHERE departament = ? ORDER BY nume", (departament,))
    users = c.fetchall()

    wb = Workbook()
    for user in users:
        user_name = user['nume']
        c.execute("SELECT data, ora_intrare, ora_iesire FROM pontaj WHERE nume = ? AND departament = ? ORDER BY data",
                  (user_name, departament))
        records = c.fetchall()
        if records:
            ws = wb.create_sheet(title=user_name)
            ws.append(["Data", "Intrare", "Iesire"])
            for record in records:
                formatted_record = (
                datetime.strptime(record['data'], '%Y-%m-%d').strftime('%d-%m-%Y'), record['ora_intrare'],
                record['ora_iesire'])
                ws.append(formatted_record)

    conn.close()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    filepath = 'pontaj_export.xlsx'
    wb.save(filepath)
    flash("Datele au fost exportate cu succes în Excel!")
    return redirect(url_for('viewer'))


if __name__ == '__main__':
    app.run(debug=True)
